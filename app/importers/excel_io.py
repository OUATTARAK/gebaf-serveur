"""Lecture Excel/CSV : extraction de lignes structurées avec détection automatique des en-têtes.

Cibles supportées :
- devis_lines : lignes de devis/facture (description, quantité, unité, PU HT)
- chantiers    : liste de chantiers (référence, nom, client, adresse, dates, budget, statut)
- clients      : liste de clients (nom, type, email, téléphone, adresse, SIRET)
"""
from __future__ import annotations
from pathlib import Path
from datetime import date, datetime
import csv
import re
import unicodedata
from typing import Any


# -------- Helpers --------

def _norm(s: Any) -> str:
    if s is None:
        return ""
    txt = str(s).strip().lower()
    # retire les accents
    txt = "".join(c for c in unicodedata.normalize("NFD", txt) if unicodedata.category(c) != "Mn")
    return re.sub(r"[\s_\-\.\(\)]+", "", txt)


def _read_rows(path: Path) -> list[list[Any]]:
    """Renvoie une liste de lignes (liste de valeurs) depuis xlsx ou csv."""
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        from openpyxl import load_workbook
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for r in ws.iter_rows(values_only=True):
            # ignorer ligne complètement vide
            if any(c is not None and str(c).strip() != "" for c in r):
                rows.append(list(r))
        wb.close()
        return rows
    if ext == ".csv":
        # auto-détection séparateur
        with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
            sample = f.read(4096)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            f.seek(0)
            reader = csv.reader(f, dialect)
            return [r for r in reader if any(c.strip() for c in r)]
    raise ValueError(f"Format de fichier non supporté : {ext}")


def _find_header_row(rows: list[list[Any]], keywords: list[tuple[str, set[str]]]) -> tuple[int, dict[str, int]]:
    """Cherche la première ligne qui contient au moins 2 mots-clés attendus.

    keywords : liste de (nom_canonique, set_de_variantes).
    Retourne (index_ligne, mapping {nom_canonique -> index}).
    """
    for i, row in enumerate(rows[:20]):
        normalized = [_norm(c) for c in row]
        found = {}
        for canonical, variants in keywords:
            for idx, cell in enumerate(normalized):
                if cell in variants:
                    found[canonical] = idx
                    break
        if len(found) >= 2:
            return i, found
    return -1, {}


def _to_float(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "").replace("\xa0", "").replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def _to_date(v: Any):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# -------- Extracteurs --------

def extract_devis_lines(path: Path) -> dict:
    """Cherche lignes devis : description, quantité, unité, PU HT."""
    rows = _read_rows(path)
    header_kw = [
        ("description", {"description", "designation", "intitule", "libelle", "produit"}),
        ("quantite", {"quantite", "qte", "quantity", "qty", "nombre"}),
        ("unite", {"unite", "unit", "u"}),
        ("prixunitaire", {"prixunitaire", "pu", "puht", "prixu", "prix", "unitprice", "tarif"}),
    ]
    hdr_idx, cols = _find_header_row(rows, header_kw)
    warnings = []
    items = []

    if hdr_idx < 0:
        # aucune en-tête trouvée — on suppose colonnes A=desc, B=qte, C=unite, D=pu
        warnings.append("En-têtes non trouvés — colonnes supposées : A=description, B=quantité, C=unité, D=PU HT")
        start = 0
        cols = {"description": 0, "quantite": 1, "unite": 2, "prixunitaire": 3}
    else:
        start = hdr_idx + 1

    for i, row in enumerate(rows[start:], start=start + 1):
        desc = str(row[cols["description"]]).strip() if "description" in cols and cols["description"] < len(row) and row[cols["description"]] else ""
        if not desc:
            continue
        qty = _to_float(row[cols["quantite"]] if "quantite" in cols and cols["quantite"] < len(row) else 1)
        unit = str(row[cols["unite"]] if "unite" in cols and cols["unite"] < len(row) and row[cols["unite"]] else "u").strip()
        pu = _to_float(row[cols["prixunitaire"]] if "prixunitaire" in cols and cols["prixunitaire"] < len(row) else 0)
        items.append({
            "description": desc,
            "quantity": qty or 1.0,
            "unit": unit or "u",
            "unit_price_ht": pu,
        })

    return {
        "kind": "devis_lines",
        "data": {"items": items},
        "warnings": warnings,
    }


def extract_chantiers(path: Path) -> dict:
    rows = _read_rows(path)
    header_kw = [
        ("reference", {"reference", "ref", "numero", "code"}),
        ("nom", {"nom", "name", "intitule", "chantier"}),
        ("client", {"client", "clientname"}),
        ("adresse", {"adresse", "address", "lieu"}),
        ("datedebut", {"datedebut", "debut", "startdate", "start", "datedepart"}),
        ("datefin", {"datefin", "fin", "enddate", "end"}),
        ("budget", {"budget", "budgetht", "montant"}),
        ("statut", {"statut", "status", "etat"}),
    ]
    hdr_idx, cols = _find_header_row(rows, header_kw)
    warnings = []
    if hdr_idx < 0:
        return {"kind": "chantiers", "data": {"rows": []}, "warnings": ["Impossible de détecter les en-têtes — ajoutez une ligne d'en-tête (nom, client, adresse, début, fin, budget, statut)"]}

    out = []
    for row in rows[hdr_idx + 1:]:
        def cell(name):
            if name in cols and cols[name] < len(row):
                v = row[cols[name]]
                return v if v is not None else ""
            return ""
        name = str(cell("nom")).strip()
        if not name:
            continue
        out.append({
            "reference": str(cell("reference")).strip() or None,
            "name": name,
            "client_name": str(cell("client")).strip() or None,
            "address": str(cell("adresse")).strip() or None,
            "start_date": _to_date(cell("datedebut")),
            "end_date": _to_date(cell("datefin")),
            "budget_ht": _to_float(cell("budget")),
            "status_raw": str(cell("statut")).strip().lower() or None,
        })
    return {"kind": "chantiers", "data": {"rows": out}, "warnings": warnings}


def extract_clients(path: Path) -> dict:
    rows = _read_rows(path)
    header_kw = [
        ("nom", {"nom", "name", "client", "raison", "raisonsociale"}),
        ("type", {"type", "typeclient"}),
        ("email", {"email", "mail", "courriel"}),
        ("telephone", {"telephone", "tel", "phone", "portable"}),
        ("adresse", {"adresse", "address"}),
        ("siret", {"siret", "siren", "rcs"}),
        ("notes", {"notes", "remarques"}),
    ]
    hdr_idx, cols = _find_header_row(rows, header_kw)
    if hdr_idx < 0:
        return {"kind": "clients", "data": {"rows": []}, "warnings": ["Aucun en-tête détecté — ajoutez une ligne d'en-tête (nom, email, téléphone, adresse, ...)"]}

    out = []
    for row in rows[hdr_idx + 1:]:
        def cell(name):
            if name in cols and cols[name] < len(row):
                v = row[cols[name]]
                return v if v is not None else ""
            return ""
        name = str(cell("nom")).strip()
        if not name:
            continue
        out.append({
            "name": name,
            "type_raw": str(cell("type")).strip().lower() or "particulier",
            "email": str(cell("email")).strip() or None,
            "phone": str(cell("telephone")).strip() or None,
            "address": str(cell("adresse")).strip() or None,
            "siret": str(cell("siret")).strip() or None,
            "notes": str(cell("notes")).strip() or None,
        })
    return {"kind": "clients", "data": {"rows": out}, "warnings": []}
